import openpyxl
from question import Question
from random import randint


def get_question_and_answer():
    questions = []
    workbook_obj = openpyxl.load_workbook("MCQ.xlsx")
    sheet_obj = workbook_obj.active
    max_row = sheet_obj.max_row

    for row_num in range(2, max_row + 1):
        question = sheet_obj.cell(row=row_num, column=1).value
        answer = sheet_obj.cell(row=row_num, column=2).value
        current_question_obj = Question(question, answer)
        questions.append(current_question_obj)

    return questions


def start_test_and_give_score():
    questions = get_question_and_answer()
    score = 0
    for turn in range(3):
        random_question_index = randint(0, len(questions) - 1)
        selected_question = questions[random_question_index]
        user_answer = input(selected_question.question)
        if user_answer == selected_question.answer:
            score += 1
    return score
